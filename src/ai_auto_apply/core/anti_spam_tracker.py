"""
Anti-Spam Tracker

Atomic Excel updater for preventing duplicate job applications.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from threading import Lock
import re
from src.common.logger import get_logger

logger = get_logger("anti_spam_tracker")

# Regex pattern for illegal Excel characters (control characters except tab, newline, carriage return)
ILLEGAL_CHARACTERS_RE = re.compile(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F]')

# Required columns for the tracker to work
_REQUIRED_COLUMNS = {"Applied", "Application Date", "Notes"}


class AntiSpamTracker:
    """Atomic Excel updater for anti-spam tracking"""
    
    def __init__(self, excel_path: str):
        """
        Initialize anti-spam tracker.
        
        Args:
            excel_path: Path to master Excel file
        """
        self.excel_path = Path(excel_path)
        self.lock = Lock()  # Thread-safe lock for atomic writes
        
        logger.info("AntiSpamTracker initialized: %s", excel_path)
    
    def _sanitize_text(self, text: str) -> str:
        """
        Sanitize text to remove illegal Excel characters.
        
        Args:
            text: Input text
            
        Returns:
            Sanitized text safe for Excel
        """
        if not isinstance(text, str):
            text = str(text)
        # Remove illegal characters
        return ILLEGAL_CHARACTERS_RE.sub('', text)
    
    def _load_dataframe(self) -> tuple:
        """
        Load Excel into DataFrame, detecting the correct sheet name.
        
        Returns:
            Tuple of (DataFrame, sheet_name)
            
        Raises:
            FileNotFoundError: If Excel file doesn't exist
        """
        # Try "Jobs" sheet first, fall back to first sheet
        try:
            df = pd.read_excel(self.excel_path, sheet_name="Jobs", engine="openpyxl")
            sheet_name = "Jobs"
        except ValueError:
            # "Jobs" sheet doesn't exist, use first sheet
            df = pd.read_excel(self.excel_path, engine="openpyxl")
            sheet_name = None  # Will use default sheet name
        
        # Ensure required columns exist (create if missing)
        for col in _REQUIRED_COLUMNS:
            if col not in df.columns:
                df[col] = ""
                logger.warning("Column '%s' was missing from Excel, created it.", col)
        
        return df, sheet_name
    
    def _save_dataframe(self, df: pd.DataFrame, sheet_name: str):
        """
        Save DataFrame back to Excel, preserving other sheets.
        
        Uses openpyxl to load the full workbook, replace only the target sheet,
        and save — keeping all other sheets intact.
        
        Args:
            df: DataFrame to save
            sheet_name: Target sheet name (None for the first/default sheet)
        """
        from openpyxl import load_workbook
        import tempfile
        import shutil
        
        # Determine actual sheet name
        wb = load_workbook(self.excel_path)
        target_sheet = sheet_name if sheet_name else wb.sheetnames[0]
        
        # Remove the old sheet and recreate it with new data
        if target_sheet in wb.sheetnames:
            sheet_index = wb.sheetnames.index(target_sheet)
            del wb[target_sheet]
        else:
            sheet_index = len(wb.sheetnames)
        
        # Create new sheet at the same position
        ws = wb.create_sheet(title=target_sheet, index=sheet_index)
        
        # Write header row
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        
        # Write data rows
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                if pd.isna(value):
                    ws.cell(row=row_idx, column=col_idx, value="")
                else:
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Atomic save: write to temp file then rename
        temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsx", dir=self.excel_path.parent)
        os.close(temp_fd)
        
        try:
            wb.save(temp_path)
            wb.close()
            # Replace original file with temporary one
            shutil.move(temp_path, self.excel_path)
            logger.debug("Successfully saved Excel file atomically: %s", self.excel_path)
        except Exception as e:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            logger.error("Failed to save Excel file: %s", e)
            raise
    
    def mark_applied(self, excel_index: int, status: str, notes: str):
        """
        Mark a job as applied with success status.
        
        Args:
            excel_index: Row index in Excel DataFrame
            status: "success" or other status
            notes: Notes to add (AI summary)
        """
        with self.lock:
            try:
                df, sheet_name = self._load_dataframe()
                
                # Validate index
                if excel_index < 0 or excel_index >= len(df):
                    logger.error("Invalid excel_index %d (DataFrame has %d rows)", excel_index, len(df))
                    return
                
                # Sanitize notes to remove illegal characters
                sanitized_notes = self._sanitize_text(notes)
                
                # Update row
                df.loc[excel_index, "Applied"] = "AI-Applied"
                df.loc[excel_index, "Application Date"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                df.loc[excel_index, "Notes"] = sanitized_notes
                
                # Save preserving all other sheets
                self._save_dataframe(df, sheet_name)
                
                logger.info("Marked job %d as AI-Applied", excel_index)
            
            except Exception as e:
                logger.error("Failed to mark job %d as applied: %s", excel_index, e)
                raise
    
    def mark_failed(self, excel_index: int, reason: str):
        """
        Mark a job application as failed.
        
        Args:
            excel_index: Row index in Excel DataFrame
            reason: Failure reason
        """
        with self.lock:
            try:
                df, sheet_name = self._load_dataframe()
                
                # Validate index
                if excel_index < 0 or excel_index >= len(df):
                    logger.error("Invalid excel_index %d (DataFrame has %d rows)", excel_index, len(df))
                    return
                
                # Sanitize reason to remove illegal characters
                sanitized_reason = self._sanitize_text(reason)
                
                # Update Notes only (keep Applied as-is)
                current_notes = df.at[excel_index, "Notes"]
                failure_note = f"[FAILED] {sanitized_reason}"
                
                if pd.isna(current_notes) or current_notes == "":
                    df.loc[excel_index, "Notes"] = failure_note
                else:
                    df.loc[excel_index, "Notes"] = f"{current_notes} | {failure_note}"
                
                # Save preserving all other sheets
                self._save_dataframe(df, sheet_name)
                
                logger.info("Marked job %d as failed: %s", excel_index, sanitized_reason)
            
            except Exception as e:
                logger.error("Failed to mark job %d as failed: %s", excel_index, e)
                raise
