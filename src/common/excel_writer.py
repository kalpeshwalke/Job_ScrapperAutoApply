"""
Excel writer module — handles reading, writing, deduplication, and formatting
of the master Excel file. Includes application-tracking columns with explicit
defaults and data-validation dropdowns.
"""

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os
import tempfile

from src.common.logger import get_logger

logger = get_logger("excel_writer")

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------
# Order matters — this is the column order in the Excel sheet.
COLUMNS = [
    "Job Title",
    "Company",
    "Company Career Page",  # Career page link
    "Career_Page_Valid",  # Career page validation status (Yes/No/Unchecked)
    "Location",
    "Experience",
    "Skills",
    "Salary",
    "Description",
    "Full Description",
    "Posted Date",
    "Platform",  # Naukri, LinkedIn, etc. (for future integrations)
    "Link",
    "Scraped Date",
    "Run Status",  # Partial vs Complete run indicator
    # --- User-editable tracking columns ---
    "Applied",
    "Application Date",
    "Status",
    "Notes",
]

# Tracking column defaults — NEVER leave blank
TRACKING_DEFAULTS = {
    "Run Status": "Complete",  # Default to Complete if not specified
    "Career_Page_Valid": "Unchecked",  # Default validation status
    "Applied": "No",
    "Application Date": "",
    "Status": "Not Applied",
    "Notes": "",
}

# Status dropdown options
STATUS_OPTIONS = "Not Applied,Applied,Interview,Rejected,Offer"


# ---------------------------------------------------------------------------
# Read existing data
# ---------------------------------------------------------------------------
def load_existing_data(file_path: Path) -> pd.DataFrame:
    """
    Load the existing master Excel file. Returns empty DataFrame if not found.
    """
    if not file_path.exists():
        logger.info("No existing master file found at %s — will create new", file_path)
        return pd.DataFrame(columns=COLUMNS)

    try:
        df = pd.read_excel(file_path, sheet_name="Jobs", engine="openpyxl")
        logger.info("Loaded %d existing jobs from %s", len(df), file_path.name)
        return df
    except Exception as e:
        logger.error("Failed to read existing master file: %s", e)
        return pd.DataFrame(columns=COLUMNS)


def get_known_links(file_path: Path) -> set:
    """
    Extract all known job links from the existing master file.
    Used for 'only_new_since_last_run' mode.
    """
    df = load_existing_data(file_path)
    if df.empty or "Link" not in df.columns:
        return set()
    links = set(df["Link"].dropna().astype(str).tolist())
    logger.info("Loaded %d known job links for fresh-only filtering", len(links))
    return links


# ---------------------------------------------------------------------------
# Deduplication
# ---------------------------------------------------------------------------
def deduplicate(new_jobs: list[dict], existing_df: pd.DataFrame) -> list[dict]:
    """
    Remove jobs that already exist in the master file.
    Dedup by: Job Link OR (Title + Company) combo.
    """
    if existing_df.empty:
        return new_jobs

    existing_links = set(existing_df["Link"].dropna().astype(str).tolist())
    existing_combos = set()
    for _, row in existing_df.iterrows():
        title = str(row.get("Job Title", "")).strip().lower()
        company = str(row.get("Company", "")).strip().lower()
        if title and company:
            existing_combos.add((title, company))

    unique_jobs = []
    dup_count = 0

    for job in new_jobs:
        link = str(job.get("link", "")).strip()
        title = str(job.get("title", "")).strip().lower()
        company = str(job.get("company", "")).strip().lower()

        # Check link-based dedup
        if link and link in existing_links:
            dup_count += 1
            continue

        # Check title+company combo dedup
        if title and company and (title, company) in existing_combos:
            dup_count += 1
            continue

        unique_jobs.append(job)
        # Add to sets to prevent intra-batch duplicates
        if link:
            existing_links.add(link)
        if title and company:
            existing_combos.add((title, company))

    logger.info("Dedup: %d new, %d duplicates removed", len(unique_jobs), dup_count)
    return unique_jobs


# ---------------------------------------------------------------------------
# Convert raw dicts to DataFrame rows
# ---------------------------------------------------------------------------
def _jobs_to_dataframe(
    jobs: list[dict],
    description_preview_chars: int = 300,
    include_full_description: bool = True,
    find_career_pages: bool = True,
    driver=None,  # Optional browser instance for Google search
    run_status: str = "Complete",  # "Complete" or "Partial"
) -> pd.DataFrame:
    """Convert raw job dicts (from scrapers) to a formatted DataFrame."""
    from src.common.company_careers import find_company_career_page
    
    rows = []
    today = datetime.now().strftime("%Y-%m-%d")
    
    # Get unique companies for career page lookup
    if find_career_pages:
        unique_companies = list(set([job.get("company", "") for job in jobs if job.get("company")]))
        logger.info("Finding career pages for %d unique companies...", len(unique_companies))
        
        # Find career pages (with progress logging and optional browser)
        career_pages = {}
        for i, company in enumerate(unique_companies):
            if i % 10 == 0 and i > 0:
                logger.info("Career page lookup progress: %d/%d", i, len(unique_companies))
            career_pages[company] = find_company_career_page(company, driver=driver)
    else:
        career_pages = {}

    for job in jobs:
        full_desc = str(job.get("description", ""))
        preview = full_desc[:description_preview_chars] if full_desc else ""
        company = job.get("company", "")

        row = {
            "Job Title": job.get("title", ""),
            "Company": company,
            "Company Career Page": career_pages.get(company, ""),
            "Career_Page_Valid": job.get("career_page_valid", "Unchecked"),
            "Location": job.get("location", ""),
            "Experience": job.get("experience", ""),
            "Skills": job.get("skills", ""),
            "Salary": job.get("salary", ""),
            "Description": preview,
            "Full Description": full_desc if include_full_description else "",
            "Posted Date": job.get("posted_date", ""),
            "Platform": job.get("platform", ""),
            "Link": job.get("link", ""),
            "Scraped Date": today,
            "Run Status": run_status,  # Partial or Complete
            # Explicit defaults for tracking columns
            "Applied": TRACKING_DEFAULTS["Applied"],
            "Application Date": TRACKING_DEFAULTS["Application Date"],
            "Status": TRACKING_DEFAULTS["Status"],
            "Notes": TRACKING_DEFAULTS["Notes"],
        }
        rows.append(row)

    return pd.DataFrame(rows, columns=COLUMNS)


# ---------------------------------------------------------------------------
# Write / Append to Excel
# ---------------------------------------------------------------------------
def save_to_excel(
    new_jobs: list[dict],
    file_path: Path,
    description_preview_chars: int = 300,
    include_full_description: bool = True,
    find_career_pages: bool = True,
    driver=None,  # Optional browser instance
    is_partial: bool = False,  # Whether this is a partial results save
    run_status: str = None,  # "Complete" or "Partial" - auto-determined if None
):
    """
    Append new jobs to the master Excel file.
    Preserves user-edited tracking columns in existing rows.
    Creates the file if it doesn't exist.
    
    Args:
        new_jobs: List of job dictionaries to save
        file_path: Path to the Excel file
        description_preview_chars: Number of characters for description preview
        include_full_description: Whether to include full description column
        find_career_pages: Whether to find company career pages
        driver: Optional browser instance for career page search
        is_partial: Whether this is a partial results save (some scrapers failed)
        run_status: Run status ("Complete" or "Partial"). Auto-determined from is_partial if None.
    """
    if not new_jobs:
        logger.info("No new jobs to save")
        return

    # Auto-determine run_status if not provided
    if run_status is None:
        run_status = "Partial" if is_partial else "Complete"

    # Load existing data
    existing_df = load_existing_data(file_path)

    # Convert new jobs to DataFrame
    new_df = _jobs_to_dataframe(
        new_jobs,
        description_preview_chars=description_preview_chars,
        include_full_description=include_full_description,
        find_career_pages=find_career_pages,
        driver=driver,  # Pass browser instance
        run_status=run_status,  # Pass run status
    )

    # Combine: existing (with user edits preserved) + new
    if not existing_df.empty:
        # Ensure existing_df has all columns (backward compat)
        for col in COLUMNS:
            if col not in existing_df.columns:
                existing_df[col] = TRACKING_DEFAULTS.get(col, "")
        combined_df = pd.concat([existing_df, new_df], ignore_index=True)
    else:
        combined_df = new_df

    # Fill any NaN in tracking columns with defaults
    for col, default in TRACKING_DEFAULTS.items():
        combined_df[col] = combined_df[col].fillna(default)
        # Also replace empty strings if needed
        if default:
            combined_df[col] = combined_df[col].replace("", default)

    # Write to Excel
    _write_formatted_excel(combined_df, file_path)
    logger.info("Saved %d new jobs to %s (total: %d)", len(new_jobs), file_path.name, len(combined_df))


# ---------------------------------------------------------------------------
# Formatted Excel writer
# ---------------------------------------------------------------------------
def _write_formatted_excel(df: pd.DataFrame, file_path: Path):
    """Write DataFrame to Excel with professional formatting."""
    file_path.parent.mkdir(parents=True, exist_ok=True)

    # Create a temporary file in the same directory to ensure atomic replacement
    temp_fd, temp_path_str = tempfile.mkstemp(suffix=".xlsx", dir=file_path.parent)
    os.close(temp_fd)
    temp_path = Path(temp_path_str)

    # Write raw data first
    with pd.ExcelWriter(temp_path, engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, sheet_name="Jobs", index=False)

    # Open for formatting
    wb = load_workbook(temp_path)
    ws = wb["Jobs"]

    # --- Header styling ---
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    tracking_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    tracking_cols = {"Applied", "Application Date", "Status", "Notes"}

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = tracking_fill if col_name in tracking_cols else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # --- Auto-adjust column widths ---
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)

        # Calculate max width from data
        max_len = len(col_name)
        for row_idx in range(2, min(ws.max_row + 1, 52)):  # Sample first 50 rows
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_len = max(max_len, min(len(str(cell_value)), 50))

        # Set width (cap at 50, min 12)
        ws.column_dimensions[col_letter].width = max(min(max_len + 3, 50), 12)

    # --- Hide "Full Description" column ---
    if "Full Description" in COLUMNS:
        fd_idx = COLUMNS.index("Full Description") + 1
        fd_letter = get_column_letter(fd_idx)
        ws.column_dimensions[fd_letter].hidden = True

    # --- Freeze top row ---
    ws.freeze_panes = "A2"

    # --- Make Link column clickable ---
    if "Link" in COLUMNS:
        link_col = COLUMNS.index("Link") + 1
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=link_col)
            if cell.value and str(cell.value).startswith("http"):
                cell.hyperlink = str(cell.value)
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

    # --- Make Company Career Page column clickable ---
    if "Company Career Page" in COLUMNS:
        career_col = COLUMNS.index("Company Career Page") + 1
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=career_col)
            if cell.value and str(cell.value).startswith("http"):
                cell.hyperlink = str(cell.value)
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

    # --- Data validation: Status dropdown ---
    if "Status" in COLUMNS:
        status_col = COLUMNS.index("Status") + 1
        status_letter = get_column_letter(status_col)
        dv = DataValidation(
            type="list",
            formula1=f'"{STATUS_OPTIONS}"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.error = "Please select a valid status"
        dv.errorTitle = "Invalid Status"
        dv.prompt = "Select application status"
        dv.promptTitle = "Status"
        dv_range = f"{status_letter}2:{status_letter}{max(ws.max_row, 1000)}"
        dv.add(dv_range)
        ws.add_data_validation(dv)

    # --- Data validation: Applied dropdown (Yes/No/AI-Applied) ---
    if "Applied" in COLUMNS:
        applied_col = COLUMNS.index("Applied") + 1
        applied_letter = get_column_letter(applied_col)
        dv_applied = DataValidation(
            type="list",
            formula1='"Yes,No,AI-Applied"',
            allow_blank=False,
            showDropDown=False,
        )
        dv_range = f"{applied_letter}2:{applied_letter}{max(ws.max_row, 1000)}"
        dv_applied.add(dv_range)
        ws.add_data_validation(dv_applied)
    
    # --- Data validation: Career_Page_Valid dropdown (Yes/No/Unchecked) ---
    if "Career_Page_Valid" in COLUMNS:
        valid_col = COLUMNS.index("Career_Page_Valid") + 1
        valid_letter = get_column_letter(valid_col)
        dv_valid = DataValidation(
            type="list",
            formula1='"Yes,No,Unchecked"',
            allow_blank=False,
            showDropDown=False,
        )
        dv_range = f"{valid_letter}2:{valid_letter}{max(ws.max_row, 1000)}"
        dv_valid.add(dv_range)
        ws.add_data_validation(dv_valid)

    # --- Conditional formatting: Applied = Yes → green fill ---
    from openpyxl.formatting.rule import CellIsRule
    if "Applied" in COLUMNS:
        applied_col = COLUMNS.index("Applied") + 1
        applied_letter = get_column_letter(applied_col)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(color="006100")
        ws.conditional_formatting.add(
            f"{applied_letter}2:{applied_letter}{max(ws.max_row, 1000)}",
            CellIsRule(operator="equal", formula=['"Yes"'], fill=green_fill, font=green_font),
        )

    # --- Cell formatting for data rows ---
    data_font = Font(name="Calibri", size=10)
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.font == Font():  # Don't override hyperlink font
                cell.font = data_font
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = thin_border

    wb.save(temp_path)
    
    # Atomic replacement
    try:
        os.replace(temp_path, file_path)
    except Exception as e:
        logger.error("Failed to replace master file with temporary file: %s", e)
        if temp_path.exists():
            logger.info("Temporary file preserved at: %s", temp_path)
        raise



# ---------------------------------------------------------------------------
# Save with career page validation
# ---------------------------------------------------------------------------
def save_to_excel_with_validation(
    new_jobs: list[dict],
    file_path: Path,
    validator,  # CareerPageValidator instance
    description_preview_chars: int = 300,
    include_full_description: bool = True,
    find_career_pages: bool = True,
    driver=None,
    is_partial: bool = False,
    run_status: str = None,
):
    """
    Save jobs to Excel with career page validation.
    
    Args:
        new_jobs: List of job dictionaries
        file_path: Path to Excel file
        validator: CareerPageValidator instance
        description_preview_chars: Number of characters for description preview
        include_full_description: Whether to include full description column
        find_career_pages: Whether to find company career pages
        driver: Optional browser instance for career page search
        is_partial: Whether this is a partial results save
        run_status: Run status ("Complete" or "Partial")
    """
    if not new_jobs:
        logger.info("No new jobs to save")
        return
    
    # Validate all career pages
    validation_results = validator.validate_batch(new_jobs)
    
    # Add validation results to job dictionaries
    for job in new_jobs:
        career_url = job.get("career_page_url", "")
        if career_url in validation_results:
            status, reason = validation_results[career_url]
            job["career_page_valid"] = status
            job["validation_reason"] = reason
        else:
            job["career_page_valid"] = "Unchecked"
            job["validation_reason"] = "Not validated"
    
    # Save to Excel with validation columns
    save_to_excel(
        new_jobs,
        file_path,
        description_preview_chars=description_preview_chars,
        include_full_description=include_full_description,
        find_career_pages=find_career_pages,
        driver=driver,
        is_partial=is_partial,
        run_status=run_status,
    )
